import asyncio

import openpyxl


async def check_values_in_excel(file_path="tables/Коэффициенты складов.xlsx", value_col3="210001", value_col6='0'):
    # Открываем .xlsx файл
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active  # Берем активный лист (или можно указать по имени, если нужно)

    # Проходим по строкам начиная с первой строки (или с другой, если есть заголовки)
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=6):
        # Проверяем значение в 3-м столбце
        if str(row[2].value) == value_col3:
            # Если соответствует, проверяем значение в 6-м столбце
            if str(row[5].value) == value_col6:
                print(True)
                return True  # Если оба условия выполняются, возвращаем True
    print(False)
    return False  # Если цикл завершился и соответствия не было найдено

asyncio.run(check_values_in_excel())
