# import time
# import pandas as pd
from loguru import logger
from openpyxl import load_workbook

# import gspread


async def data_to_exel(file_name, headers, data, headers_rus=None):
    try:
        wb = load_workbook(file_name)
        ws = wb.active
        # print(wb.sheetnames)
        ws.delete_rows(1, ws.max_row)
        # print(headers)
        if headers_rus is None:
            headers_rus = headers
        for col, header in enumerate(headers_rus, start=1):
            ws.cell(row=1, column=col, value=header)
        # Записываем данные
        for row, row_data in enumerate(data, start=2):
            for col, key in enumerate(headers, start=1):
                ws.cell(row=row, column=col, value=row_data[key])
        # Сохраняем файл
        wb.save(file_name)
    except Exception as e:
        logger.exception("Ошибка записи в Excel", exc_info=e)


# data_to_exel("box rates.xlsx", [{'boxDeliveryAndStorageExpr': '205', 'boxDeliveryBase': '61,5', 'boxDeliveryLiter': '14,35', 'boxStorageBase': '0,328', 'boxStorageLiter': '0,287', 'warehouseName': 'Коледино'}, {'boxDeliveryAndStorageExpr': '280', 'boxDeliveryBase': '84', 'boxDeliveryLiter': '19,6', 'boxStorageBase': '0,448', 'boxStorageLiter': '0,392', 'warehouseName': 'Подольск'}])
